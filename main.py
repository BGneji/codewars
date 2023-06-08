from openpyxl import Workbook, load_workbook
import datetime
import os


# создание Workbook
workbook = Workbook()

active_worksheet = workbook.active
active_worksheet['A1'] = 'Hello World'
active_worksheet['B1'] = datetime.datetime.now()
workbook.save('test_worbook.xlsx')
print(os.listdir('.'))

workbook = load_workbook(filename='test_workbook.xlsx')
print('Sheetname are: ' + str(workbook.sheetnames))
sheet = workbook.active
print('Value of cell A1: ' + sheet['A1'].value)

print(sheet['A:B'])
print(sheet[1])